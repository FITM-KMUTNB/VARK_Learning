from varkapp.models import Content, User, Topic, Exercise, Chapter
from flask_login import current_user

def topic_summary():
    userid = User.query.filter_by(email='5906021610078@fitm.kmutnb.ac.th').first().id
    exerciseDB = Exercise.query.filter_by(user_id=userid).all()

    user_exercise = dict() # {Chapter : {Topic : {Learntype : percent } } }
    for ex in exerciseDB:
        db_topic = Topic.query.filter_by(id=ex.topic_id).first()
        
        #print('Chapter : ', db_topic.chapter_id)
        if db_topic.chapter_id not in user_exercise:
            user_exercise[db_topic.chapter_id] = {}
        #print('Topic : ', db_topic.number)
        if db_topic.number not in user_exercise[db_topic.chapter_id]:
            user_exercise[db_topic.chapter_id][db_topic.number] = {}

        #print("Learn :", ex.learntype, " ,Percent : ", ex.percent)
        if ex.learntype not in user_exercise[db_topic.chapter_id][db_topic.number]:
            user_exercise[db_topic.chapter_id][db_topic.number][ ex.learntype] = ex.percent

    print(user_exercise)

#topic_summary()
from openpyxl import Workbook

def print_out_report():
    workbook = Workbook()
    sheet = workbook.active

    # user column
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=5)
    sheet["A1"] = "User"
    sheet["A3"] = "Firstname"
    sheet["B3"] = "Lastname"
    sheet["C3"] = "Gender"
    sheet["D3"] = "Age"
    sheet["E3"] = "Email"

    # learn column
    start_column = 6
    topic_row = 2
    topic_start_column = 6
    content_row = 3
    allchapter = Chapter.query.all()

    for chapter in allchapter:
    
        chapter_no = chapter.number
        chapter_topic = Topic.query.filter_by(chapter_id=chapter.id).all()

        chapter_merge = 0
        #print("chapter : ", chapter.number)
        for topic in chapter_topic:
            topic_merge = 1
            if topic.number != 'P' and topic.number != 'T':
                chapter_merge += 6
                topic_merge = 6
                content_start_column = topic_start_column 
                for media in ['V', 'A', 'R', 'K','Times', 'Score']:
                    sheet.cell(row=content_row, column=content_start_column).value = media
                    content_start_column += 1
            else:
                chapter_merge += 1
                topic_merge = 1
                sheet.cell(row=content_row, column=topic_start_column).value = topic.number

            #print(topic.number, ", topic merge: ", topic_merge)
            topic_end_column = topic_start_column + topic_merge - 1
            sheet.merge_cells(start_row=topic_row, start_column=topic_start_column, end_row=topic_row, end_column=topic_end_column)
            sheet.cell(row=topic_row, column=topic_start_column).value = topic.number
            topic_start_column = topic_end_column + 1
        
        #print("chapter merge : ", chapter_merge)
        chapter_row = 1
        end_column = start_column + chapter_merge - 1
        sheet.merge_cells(start_row=chapter_row, start_column=start_column, end_row=chapter_row, end_column=end_column)
        sheet.cell(row=chapter_row, column=start_column).value = chapter.name

        start_column = end_column + 1

    # user excercise
    alluser = User.query.all()
    user_row_start = 4
    for user in alluser:
        user_column = 1
        if user.user_type != 'Admin':
            firstname = user.firstname
            lastname = user.lastname
            gender = user.gender
            age = user.age
            email = user.email
    
            for uinfo in [firstname, lastname, gender, age, email]:
                sheet.cell(row=user_row_start, column=user_column).value = uinfo
                user_column += 1
         
            for chapter in allchapter:
                chapter_topic = Topic.query.filter_by(chapter_id=chapter.id).all()
                for topic in chapter_topic:
                    user_excercise = Exercise.query.filter_by(user_id=user.id, topic_id = topic.id).first()
                    if user_excercise:
                        if topic.number == 'P' or topic.number == 'T':
                            sheet.cell(row=user_row_start, column=user_column).value = user_excercise.percent
                            user_column += 1
                        else:
                            for media in ['V', 'A', 'R', 'K']:  
                                media_point = Exercise.query.filter_by(user_id=user.id, topic_id = topic.id, learntype=media).all()
                                
                                if media_point:
                                    highest_percent = 0
                                    for mp in media_point:
                                        if highest_percent < int(float(mp.percent)):
                                            highest_percent = int(float(mp.percent))
                                    sheet.cell(row=user_row_start, column=user_column).value = highest_percent
                                    user_column += 1
                                else:
                                    sheet.cell(row=user_row_start, column=user_column).value = '-'
                                    user_column += 1

                            topic_score = Exercise.query.filter_by(user_id=user.id, topic_id = topic.id).all()
                            score_vark = ''
                            num = 0
                            for ts in topic_score:
                                score_vark += ts.learntype + ':' +str(int(float(ts.percent)))
                                num += 1
                                if num < len(topic_score):
                                    score_vark += ', '
                            sheet.cell(row=user_row_start, column=user_column).value = len(topic_score)
                            user_column += 1
                            sheet.cell(row=user_row_start, column=user_column).value = score_vark
                            user_column += 1
            user_row_start += 1
    workbook.save(filename="vark_report.xlsx")

#print_out_report()


def chapter_summary():
    chapter_sum = dict() #{1: {'V': 25, 'R': 53, 'A': 80, 'K': 50}}
    vark_earn_score = dict()
    vark_full_score = dict()
    allchapter = Chapter.query.all()
    userid = User.query.filter_by(email='5906021610078@fitm.kmutnb.ac.th').first().id

    # iterate chapter
    for ch in allchapter:
        chapter_vark = dict()
        chapter_topic = Topic.query.filter_by(chapter_id=ch.id)
        # iterate chapter topic
        for tp in chapter_topic:
            # filter topic excercise of user
            user_result = Exercise.query.filter_by(user_id=userid, topic_id=tp.id).all()
            # skip pretest and posttest
            if tp.number != 'P' and tp.number != 'T' and user_result:
                # get v a r k earned score of topic
                topic_vark = dict()
                for us in user_result:
                    if us.learntype in topic_vark:
                        # in case do the same vark sevaral time, get highest score.
                        if topic_vark[us.learntype] < int(us.getpoint):
                            topic_vark[us.learntype] = int(us.getpoint)
                    else:
                        topic_vark[us.learntype] = int(us.getpoint)
                # score this toppic score to chapter dict()
                for tk in topic_vark:
                    if tk in chapter_vark:
                        chapter_vark[tk] += topic_vark[tk]
                    else:
                        chapter_vark[tk] = topic_vark[tk]

                    # sum of full score from first chapter and keep increasing
                    if tk in vark_full_score:
                        vark_full_score[tk] += 5
                    else:
                        vark_full_score[tk] = 5
                

            # if done posttest
            elif tp.number == 'T' and user_result:
                for ck in chapter_vark:
                    if ck in vark_earn_score:
                        vark_earn_score[ck] += chapter_vark[ck]
                    else:
                        vark_earn_score[ck] = chapter_vark[ck]
                
                chapter_sum[ch.id] = {}
                for vs in vark_earn_score:
                    chapter_sum[ch.id][vs] = int((vark_earn_score[vs] / vark_full_score[vs]) * 100)

    return chapter_sum
               
#chapter_summary()


