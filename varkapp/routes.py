from flask import render_template, url_for, request, flash, redirect, jsonify, send_file
from varkapp import app, db, bcrypt
from varkapp.forms import RegistrationForm, LoginForm
from varkapp.models import get_content, User, Topic, Exercise, Content, Chapter
from flask_login import login_user, current_user, logout_user, login_required
import pdfplumber
import glob
from datetime import datetime


@app.route('/', methods=['GET', 'POST'])
@app.route('/index', methods=['GET', 'POST'])

@login_required
def index():
    chapter, topic, content = get_content() 
    userid = User.query.filter_by(email=current_user.email).first().id
    exerciseDB = Exercise.query.filter_by(user_id=userid).all()
   
    chapter_sum = chapter_summary()

    user_exercise = dict() # {Chapter : {Topic : {Learntype : percent } } }
    for ex in exerciseDB:
        db_topic = Topic.query.filter_by(id=ex.topic_id).first()
        
        #print('Chapter : ', db_topic.chapter_id)
        if db_topic.chapter_id not in user_exercise:
            user_exercise[db_topic.chapter_id] = {}
        #print('Topic : ', db_topic.number)
        if db_topic.number not in user_exercise[db_topic.chapter_id]:
            user_exercise[db_topic.chapter_id][db_topic.number] = {}

        #print("Learn :", ex.learntype, " ,Percent : ", ex.percent)
        if ex.learntype not in user_exercise[db_topic.chapter_id][db_topic.number]:
            user_exercise[db_topic.chapter_id][db_topic.number][ex.learntype] = int(float(ex.percent))
        else:
            del user_exercise[db_topic.chapter_id][db_topic.number][ex.learntype] 
            user_exercise[db_topic.chapter_id][db_topic.number][ex.learntype] = int(float(ex.percent))
 
    return render_template('index.html', title = "VARK", chapter=chapter, topic=topic, content=content,\
    user_exercise=user_exercise, User=User, Topic=Topic, chapter_sum = chapter_sum)

@app.route('/registration', methods=['GET', 'POST'])
def registration():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    title = "Registration"
    form = RegistrationForm()
    if form.validate_on_submit():
        hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        user = User(gender=form.gender.data, firstname=form.firstname.data, 
                    lastname=form.lastname.data, age=form.age.data, email=form.email.data, password=hashed_password, user_type='User1')
        db.session.add(user)
        db.session.commit()
        flash('Your account has been created! You are now able to log in', 'success')
        return redirect(url_for('login'))
    return render_template('registration.html', title = title, form=form)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    title = "Sign In"
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user and bcrypt.check_password_hash(user.password, form.password.data):
            login_user(user, remember=form.remember.data)
            return redirect(url_for('index'))
        else:
            flash('Login Unsuccessful. Please check email and password', 'danger')

    return render_template('login.html', title=title, form=form)

@app.route("/logout")
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/display_exercise', methods=['GET', 'POST'])
def display_exercise():
    if request.method == 'POST':
        # display excercise
        # question number for check question line in text.

        start=datetime.now()

        q_number = []
        for num in range(1,20):
            q_number.append(str(num))
            q_number.append(str(num)+'.')
            for num2 in range(1,20):
                for num3 in range(1,20):
                    q_number.append(str(num)+"."+str(num2)+"."+str(num3))

        # extract exercise file
        testfile= request.form['testfile']

        # keep chapter id and topic id for query in database.
        chapterid = request.form['chapterid']
        topicid = request.form['topicid']
        learntype = request.form['learntype']

        #pdf = pdfplumber.open("varkapp/"+testfile)
        pdf = pdfplumber.open('./'+testfile)
        print(testfile)
        text = []
        for page in range(0, len(pdf.pages)):
            pages = pdf.pages[page].extract_text().splitlines()
            for line in pages:
                
                # Replace "า" to '�' if before '�' is not 'ำ"
                if '�' in line:
                    my_new_string = ""
                    cutt = line.split("�")
                    
                    for index in range(len(cutt)):
                        if len(cutt[index]) > 0:
                            prev_char = len(cutt[index])
                            if cutt[index][prev_char-1] != "ำ" and cutt[index][prev_char-1] != " " and index != len(cutt)-1:
                                my_new_string += cutt[index]+"า"
                                
                            else:
                                my_new_string += cutt[index]
                        
                    my_new_string2 = ""
                    
                    # Remove "า" in case if "ำา" exist in my_new_string
                    if 'า' in my_new_string:
                        cutt = my_new_string.split("า")
                        
                        for index in range(len(cutt)):
                            if len(cutt[index]) > 0:
                                prev_char = len(cutt[index])
                                if cutt[index][prev_char-1] != "ำ" and cutt[index][prev_char-1] != " " and index != len(cutt)-1:
                                    my_new_string2 += cutt[index]+"า"
                                    
                                else:
                                    my_new_string2 += cutt[index]
                        text.append(my_new_string2)      
                    else:
                        text.append(my_new_string)

                # Remove "า" in case if "ำา" exist in line
                elif  'า' in line:
                    cutt = line.split("า")
                    my_new_string = "" 
                    for index in range(len(cutt)):
                        if len(cutt[index]) > 0:
                            prev_char = len(cutt[index])
                            if cutt[index][prev_char-1] != "ำ" and cutt[index][prev_char-1] != " " and index != len(cutt)-1:
                                my_new_string += cutt[index]+"า"
                                
                            else:
                                my_new_string += cutt[index]
                            
                    text.append(my_new_string)

                else:
                    text.append(line)

        #{ข้อ1 : {ก: choice, ข: choice}}
        exercise_content = dict() 
        tempt_choice = dict()
        question = None
        for t in range(len(text)):
            choice = ""
            line_cutt = text[t].split("\t")
            line_cutt = ' '.join(line_cutt).split()

            if line_cutt[0] in q_number:
                question = ""
                question += text[t]
                if t+1 < len(text):
                    check_newline = text[t+1].split("\t")
                    check_newline = ' '.join(check_newline).split()
                    if check_newline[0] not in ['ก.', 'ข.', 'ค.', 'ง.', 'ก', 'ข', 'ค', 'ง']:
                        question += '\n'+text[t+1]
                exercise_content[question] = None
            
            elif line_cutt[0] in ['ก.', 'ข.', 'ค.', 'ง.', 'ก', 'ข', 'ค', 'ง',]:
                for c in line_cutt[1:]:
                    choice += " "+c
            
                if t+1 < len(text):
                    check_newline = text[t+1].split("\t")
                    check_newline = ' '.join(check_newline).split()
                    if check_newline[0] not in ['ก.', 'ข.', 'ค.', 'ง.', 'ก', 'ข', 'ค', 'ง'] and \
                    check_newline[0] not in q_number:
                        choice += '\n'+text[t+1]
                    
                tempt_choice[line_cutt[0]] = choice
            
                if len(tempt_choice) == 4:
                    exercise_content[question] = tempt_choice
                    tempt_choice = dict()

        return render_template('exercise.html', content = exercise_content, title = text[0], testfile=testfile,\
                                chapterid=chapterid, topicid=topicid, learntype=learntype, start_time = start)

@app.route('/submit_exercise', methods=['GET', 'POST'])
def submit_exercise():

    #receive answer from template
    if request.method == 'POST':
        #print("Submit")
        # Request get from template
        quiz_number = request.form.get('quiz_number')
        testfile = request.form.get('testfile')
        chapterid = request.form.get('chapterid')
        topicid = request.form.get('topicid')
        learntype = request.form.get('learntype')
        # exercise time
        str_start_time = request.form.get('start_time')
        start_time = datetime.strptime(str_start_time, '%Y-%m-%d %H:%M:%S.%f')
        time_delta = (datetime.now() - start_time)
        total_seconds = time_delta.total_seconds()
        minutes = total_seconds/60

        exercise_time = 0
        if minutes < 1:
            exercise_time = 1
        else:
            exercise_time = minutes

        if learntype == "":
                learntype = "-"

        select_choice = []
        for select in range(1, int(quiz_number)+1):
            select_choice.append(request.form['choice'+str(select)].replace(".",""))
        
        # get anwser file path
        file_path = testfile.split("/")
        del file_path[-1]
        answer_file = "varkapp/"
        for p in range(len(file_path)):
            answer_file += file_path[p]+"/"
            
        for file in glob.glob(answer_file+"/*.txt"):
            answer_file = file

        # read answer in text file
        Text_file = open(answer_file, 'r', encoding='utf-8-sig')
        #Text_file = open(answer_file, 'r', encoding='utf-8')
        answer_choice = []
        for check_c in Text_file:
            
            check_c = check_c.replace("\n","")
         
            if len(check_c.split()) > 1:
                ans, topic = check_c.split()
            else:
                ans = check_c
            #print(ans)
            answer_choice.append(ans)
        
        # check score
        full_point = len(answer_choice)
        get_points = 0
        for ans in range(len(answer_choice)):
            if select_choice[ans] == answer_choice[ans]:
                get_points += 1
        
        # commit to database
        db_topicid = Topic.query.filter_by(chapter_id=chapterid, number=topicid).first().id
        userid = User.query.filter_by(email=current_user.email).first().id
        
        # Add to Exercise table
        exerciseDB = Exercise.query.filter_by(user_id=userid, topic_id=db_topicid).first()
        # add new exercise score
        percent = (get_points/full_point) * 100
        if User.query.filter_by(email=current_user.email).first().user_type == 'User':
            exerciseDB = Exercise(learntype=learntype, fullpoint=full_point, getpoint=get_points,percent=percent,\
                                    user_id=userid, topic_id=db_topicid)
        else:
            print(exercise_time)
            exerciseDB = Exercise(learntype=learntype, fullpoint=full_point, getpoint=get_points,percent=percent,\
                                    user_id=userid, topic_id=db_topicid, time=exercise_time)

        db.session.add(exerciseDB)
        db.session.commit()
        print(Topic.query.filter_by(id=db_topicid).first(), " : Submit")

        # Exercise Result
        #print("############## Result ######################")
        #print("User : ",  User.query.filter_by(email=current_user.email).first().firstname)
        #print("Select : ", select_choice)
        print("Answer : ", answer_choice)
        #print("Learn type : ", learntype)
        #print("correct : ", get_points)

        return redirect(url_for('index'))

@app.route('/vark_report')
def vark_report():
    print_out_report()
    return send_file('vark_report.xlsx',  as_attachment=True, mimetype='application/vnd.ms-excel',)

def chapter_summary():
    chapter_sum = dict() #{1: {'V': 25, 'R': 53, 'A': 80, 'K': 50}}
    vark_earn_score = dict()
    vark_full_score = dict()
    allchapter = Chapter.query.all()
    userid = User.query.filter_by(email=current_user.email).first().id

    # iterate chapter
    for ch in allchapter:
        chapter_vark = dict()
        chapter_topic = Topic.query.filter_by(chapter_id=ch.id)
        # iterate chapter topic
        for tp in chapter_topic:
            # filter topic excercise of user
            user_result = Exercise.query.filter_by(user_id=userid, topic_id=tp.id).all()
            # skip pretest and posttest
            if tp.number != 'P' and tp.number != 'T' and user_result:
                # get v a r k earned score of topic
                topic_vark = dict()
                for us in user_result:
                    if us.learntype in topic_vark:
                        # in case do the same vark sevaral time, get highest score.
                        if topic_vark[us.learntype] < int(us.getpoint):
                            topic_vark[us.learntype] = int(us.getpoint)
                    else:
                        topic_vark[us.learntype] = int(us.getpoint)
                # score this toppic score to chapter dict()
                for tk in topic_vark:
                    if tk in chapter_vark:
                        chapter_vark[tk] += topic_vark[tk]
                    else:
                        chapter_vark[tk] = topic_vark[tk]

                    # sum of full score from first chapter and keep increasing
                    if tk in vark_full_score:
                        vark_full_score[tk] += 5
                    else:
                        vark_full_score[tk] = 5
                

            # if done posttest
            elif tp.number == 'T' and user_result:
                for ck in chapter_vark:
                    if ck in vark_earn_score:
                        vark_earn_score[ck] += chapter_vark[ck]
                    else:
                        vark_earn_score[ck] = chapter_vark[ck]
                
                chapter_sum[ch.id] = {}
                for vs in vark_earn_score:
                    chapter_sum[ch.id][vs] = int((vark_earn_score[vs] / vark_full_score[vs]) * 100)

    return chapter_sum

from openpyxl import Workbook

def print_out_report():
    workbook = Workbook()
    sheet = workbook.active

    # user column
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=5)
    sheet["A1"] = "User"
    sheet["A3"] = "Firstname"
    sheet["B3"] = "Lastname"
    sheet["C3"] = "Gender"
    sheet["D3"] = "Age"
    sheet["E3"] = "Email"

    # learn column
    start_column = 6
    topic_row = 2
    topic_start_column = 6
    content_row = 3
    allchapter = Chapter.query.all()

    for chapter in allchapter:
    
        chapter_no = chapter.number
        chapter_topic = Topic.query.filter_by(chapter_id=chapter.id).all()

        chapter_merge = 0
        #print("chapter : ", chapter.number)
        for topic in chapter_topic:
            topic_merge = 1
            if topic.number != 'P' and topic.number != 'T':
                chapter_merge += 6
                topic_merge = 6
                content_start_column = topic_start_column 
                for media in ['V', 'A', 'R', 'K','Ex Count', 'All score']:
                    sheet.cell(row=content_row, column=content_start_column).value = media
                    content_start_column += 1
            else:
                chapter_merge += 2
                topic_merge = 2
                sheet.cell(row=content_row, column=topic_start_column).value = 'Score'
                sheet.cell(row=content_row, column=topic_start_column+1).value = 'Time'

            topic_end_column = topic_start_column + topic_merge - 1
            sheet.merge_cells(start_row=topic_row, start_column=topic_start_column, end_row=topic_row, end_column=topic_end_column)
            sheet.cell(row=topic_row, column=topic_start_column).value = topic.number
            topic_start_column = topic_end_column + 1
        
        #print("chapter merge : ", chapter_merge)
        chapter_row = 1
        end_column = start_column + chapter_merge - 1
        sheet.merge_cells(start_row=chapter_row, start_column=start_column, end_row=chapter_row, end_column=end_column)
        sheet.cell(row=chapter_row, column=start_column).value = chapter.name

        start_column = end_column + 1

    # user excercise
    alluser = User.query.all()
    user_row_start = 4
    for user in alluser:
        user_column = 1
        if user.user_type != 'Admin':
            firstname = user.firstname
            lastname = user.lastname
            gender = user.gender
            age = user.age
            email = user.email
    
            for uinfo in [firstname, lastname, gender, age, email]:
                sheet.cell(row=user_row_start, column=user_column).value = uinfo
                user_column += 1
         
            for chapter in allchapter:
                chapter_topic = Topic.query.filter_by(chapter_id=chapter.id).all()
                for topic in chapter_topic:
                    user_excercise = Exercise.query.filter_by(user_id=user.id, topic_id = topic.id).first()
                    if user_excercise:
                        if topic.number == 'P' or topic.number == 'T':
                            sheet.cell(row=user_row_start, column=user_column).value = user_excercise.percent
                            user_column += 1
                            # type = User, not have time stramp in exercise 
                            if user.user_type == 'User':
                                sheet.cell(row=user_row_start, column=user_column).value = '-'
                                user_column += 1
                            else:
                                sheet.cell(row=user_row_start, column=user_column).value = user_excercise.time
                                user_column += 1
                        else:
                            for media in ['V', 'A', 'R', 'K']:  
                                media_point = Exercise.query.filter_by(user_id=user.id, topic_id = topic.id, learntype=media).all()
                                
                                if media_point:
                                    highest_percent = 0
                                    for mp in media_point:
                                        if highest_percent < int(float(mp.percent)):
                                            highest_percent = int(float(mp.percent))
                                    sheet.cell(row=user_row_start, column=user_column).value = highest_percent
                                    user_column += 1
                                else:
                                    sheet.cell(row=user_row_start, column=user_column).value = '-'
                                    user_column += 1

                            topic_score = Exercise.query.filter_by(user_id=user.id, topic_id = topic.id).all()
                            score_vark = ''
                            num = 0
                            for ts in topic_score:
                                # score column
                                # User --> learntype : score, ..
                                if user.user_type == 'User':
                                    score_vark += ts.learntype + ':' +str(int(float(ts.percent)))
                                    num += 1
                                # User1 --> learntype : score : time, ..
                                else:
                                    score_vark += ts.learntype + ':' +str(int(float(ts.percent))) + ':' +str(int(float(ts.time)))
                                    num += 1

                                if num < len(topic_score):
                                    score_vark += ', '
                            sheet.cell(row=user_row_start, column=user_column).value = len(topic_score)
                            user_column += 1
                            sheet.cell(row=user_row_start, column=user_column).value = score_vark
                            user_column += 1
            user_row_start += 1
    #workbook.save(filename="varkapp/vark_report.xlsx")
    workbook.save(filename="vark_report.xlsx")

